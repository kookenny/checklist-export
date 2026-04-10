"""
checklist_extract.py
────────────────────
Extracts full checklist content from CaseWare Cloud SE author templates.

Outputs an Excel workbook with one sheet per checklist containing:
  - Full procedure hierarchy (groups, subgroups, procedures, sub-procedures)
  - Procedure text, AICPA/CAS standards, assertions tested
  - Lightbulb guidance content
  - Cloud settings (sign-offs, response types, options, etc.)
  - Visibility conditions (formatted as human-readable text)

SETUP
─────
1. Install dependencies:
       pip install requests openpyxl python-dotenv

2. Set OAuth credentials in .env (see .env.example for format).

3. Run:
       python checklist_extract.py --mock                      # test with sample data
       python checklist_extract.py --url "<caseware_url>"      # extract checklists
       python checklist_extract.py --discover --url "<url>"    # dump raw procedure JSON
"""

import argparse
import io
import json
import os
import logging
import re
from html import unescape
from typing import Optional

import requests as http_lib
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONFIGURATION ─────────────────────────────────────────────────────────────

HOST   = "https://ca.cwcloudpartner.com"
TENANT = "ca-develop"

OUTPUT_FILE = ".tmp/checklist_extract.xlsx"

# ── END CONFIGURATION ─────────────────────────────────────────────────────────

logging.basicConfig(
    level=os.environ.get("CW_LOG_LEVEL", "INFO").upper(),
    format="%(levelname)s  %(message)s",
)
log = logging.getLogger(__name__)


# ── HTML / TEXT PROCESSING ───────────────────────────────────────────────────

def strip_html(html: str, formula_map: dict[str, str] | None = None) -> str:
    """Convert an HTML string to plain text by removing tags and unescaping entities.
    Placeholder spans are wrapped in (( )).
    Formula/dynamic-text spans are resolved via *formula_map* and wrapped in [[ ]]."""
    if not html:
        return ""
    text = re.sub(
        r'<span[^>]*\bplaceholder="[^"]*"[^>]*>(.*?)</span>',
        r"((\1))",
        html,
    )
    if formula_map:
        def _resolve_formula(m: re.Match) -> str:
            fid = m.group(1)
            val = formula_map.get(fid, "")
            return f"[[{val}]]" if val else ""
        text = re.sub(
            r'<span[^>]*\bformula="([^"]*)"[^>]*>.*?</span>',
            _resolve_formula,
            text,
        )
    text = re.sub(r"<[^>]+>", " ", text)
    text = unescape(text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


# ── SESSION ───────────────────────────────────────────────────────────────────

def _env_prefix_from_host(host: str) -> str:
    """Derive an env-var prefix from the hostname: 'ca.cwcloudpartner.com' -> 'CA'."""
    hostname = host.replace("https://", "").replace("http://", "").split("/")[0]
    return hostname.split(".")[0].upper()


def _obtain_bearer_token(env_prefix: str = "",
                         host: str | None = None,
                         tenant: str | None = None) -> str | None:
    """Exchange CW_CLIENT_ID + CW_CLIENT_SECRET for a Bearer token via OAuth."""
    host = host or HOST
    tenant = tenant or TENANT
    client_id, client_secret = "", ""
    if env_prefix:
        client_id     = os.environ.get(f"CW_{env_prefix}_CLIENT_ID", "").strip()
        client_secret = os.environ.get(f"CW_{env_prefix}_CLIENT_SECRET", "").strip()
    if not client_id or not client_secret:
        client_id     = os.environ.get("CW_CLIENT_ID", "").strip()
        client_secret = os.environ.get("CW_CLIENT_SECRET", "").strip()
    if not client_id or not client_secret:
        return None
    url = f"{host}/{tenant}/ms/caseware-cloud/api/v1/auth/token"
    resp = http_lib.post(url, json={
        "ClientId": client_id, "ClientSecret": client_secret, "Language": "en",
    }, headers={"Accept": "application/json", "Content-Type": "application/json"},
       timeout=15)
    resp.raise_for_status()
    token = resp.json().get("Token")
    if token:
        log.info("Authenticated via OAuth (Bearer token)")
    return token


def make_session(env_prefix: str = "",
                 host: str | None = None,
                 tenant: str | None = None) -> http_lib.Session:
    """Build a requests.Session using OAuth (preferred) or browser cookies."""
    session = http_lib.Session()
    session.headers.update({
        "Accept":       "application/json",
        "Content-Type": "application/json",
    })
    token = _obtain_bearer_token(env_prefix, host=host, tenant=tenant)
    if token:
        session.headers["Authorization"] = f"Bearer {token}"
        return session
    cookies = os.environ.get("CW_COOKIES", "").strip()
    if not cookies:
        raise RuntimeError(
            "No auth credentials found. "
            "Set CW_CLIENT_ID + CW_CLIENT_SECRET for OAuth, "
            "or CW_COOKIES for cookie auth."
        )
    session.headers["Cookie"] = cookies
    log.info("Authenticated via browser cookies")
    return session


# ── API HELPERS ───────────────────────────────────────────────────────────────

def _unwrap_response(data) -> list[dict]:
    """Extract a list of objects from the various API response wrappers."""
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        for key in ("objects", "sections", "items", "results", "data"):
            if key in data and isinstance(data[key], list):
                return data[key]
        for key, val in data.items():
            if isinstance(val, list):
                return val
        if "object" in data and isinstance(data["object"], dict):
            return [data["object"]]
    return []


def _api_post(session: http_lib.Session, url: str, payload: dict,
              timeout: int = 30) -> list[dict]:
    """POST to a Caseware API endpoint and return the unwrapped list."""
    resp = session.post(url, json=payload, timeout=timeout)
    if resp.status_code == 401:
        raise RuntimeError("401 Unauthorised — credentials have expired.")
    if not resp.ok:
        raise RuntimeError(f"{resp.status_code} from {url}\n{resp.text[:500]}")
    return _unwrap_response(resp.json())


# ── DATA FETCHING ─────────────────────────────────────────────────────────────

def fetch_documents(session: http_lib.Session,
                    engagement_id: str,
                    host: str | None = None,
                    tenant: str | None = None) -> list[dict]:
    """Fetch all documents in the engagement."""
    host = host or HOST
    tenant = tenant or TENANT
    url = f"{host}/{tenant}/e/eng/{engagement_id}/api/v1.12.0/document/get"
    log.info("Fetching documents from %s", url)
    return _api_post(session, url, {})


def fetch_procedures(session: http_lib.Session,
                     engagement_id: str,
                     checklist_id: str,
                     host: str | None = None,
                     tenant: str | None = None) -> list[dict]:
    """Fetch all procedures belonging to a checklist document."""
    host = host or HOST
    tenant = tenant or TENANT
    url = f"{host}/{tenant}/e/eng/{engagement_id}/api/v1.12.0/procedure/get"
    payload = {"filter": {"filter": {
        "node": "=",
        "left":  {"node": "field", "kind": "procedure", "field": "checklistId"},
        "right": {"node": "string", "value": checklist_id},
    }}}
    return _api_post(session, url, payload)


def fetch_procedure_by_id(session: http_lib.Session,
                          engagement_id: str,
                          procedure_id: str,
                          host: str | None = None,
                          tenant: str | None = None) -> Optional[dict]:
    """Fetch a single procedure by its ID."""
    host = host or HOST
    tenant = tenant or TENANT
    url = f"{host}/{tenant}/e/eng/{engagement_id}/api/v1.12.0/procedure/get"
    payload = {"filter": {"filter": {
        "node": "=",
        "left":  {"node": "field", "kind": "procedure", "field": "id"},
        "right": {"node": "string", "value": procedure_id},
    }}}
    try:
        resp = session.post(url, json=payload, timeout=15)
        if not resp.ok:
            return None
        objects = _unwrap_response(resp.json())
        return objects[0] if objects else None
    except Exception as exc:
        log.debug("Could not fetch procedure %s: %s", procedure_id, exc)
        return None


def fetch_checklist_name(session: http_lib.Session,
                         engagement_id: str,
                         checklist_id: str,
                         host: str | None = None,
                         tenant: str | None = None) -> str:
    """Try to resolve a checklist ID to its human-readable name.
    Tries multiple strategies in order."""
    host = host or HOST
    tenant = tenant or TENANT

    # Strategy 0: fetch as procedure by its own ID
    proc = fetch_procedure_by_id(session, engagement_id, checklist_id, host=host, tenant=tenant)
    if proc:
        name = _get_procedure_name(proc)
        if name:
            number = proc.get("number", "")
            return f"{number} {name}".strip() if number else name

    # Strategy 1a: fetch as document/checklist/workpaper
    for endpoint in ("document/get", "checklist/get", "workpaper/get"):
        url = f"{host}/{tenant}/e/eng/{engagement_id}/api/v1.12.0/{endpoint}"
        for payload in (
            {"filter": {"filter": {"node": "=",
                "left": {"node": "field", "kind": endpoint.split("/")[0], "field": "id"},
                "right": {"node": "string", "value": checklist_id}}}},
            {"id": checklist_id},
        ):
            try:
                resp = session.post(url, json=payload, timeout=10)
                if not resp.ok:
                    continue
                data = resp.json()
                objects = _unwrap_response(data)
                for obj in objects:
                    for field in ("name", "names", "title", "titles", "description"):
                        val = obj.get(field)
                        if isinstance(val, dict):
                            val = val.get("en", "") or next(iter(val.values()), "")
                        if val and isinstance(val, str):
                            return val.strip()
            except Exception:
                continue

    # Strategy 1b: fetch as section by id
    url = f"{host}/{tenant}/e/eng/{engagement_id}/api/v1.12.0/section/get"
    payload = {"filter": {"filter": {
        "node": "=",
        "left":  {"node": "field", "kind": "section", "field": "id"},
        "right": {"node": "string", "value": checklist_id},
    }}}
    try:
        resp = session.post(url, json=payload, timeout=10)
        if resp.ok:
            objects = _unwrap_response(resp.json())
            if objects:
                title = strip_html(objects[0].get("title", ""))
                if title:
                    return title
    except Exception:
        pass

    return ""


def fetch_checklist_defaults(session: http_lib.Session,
                             engagement_id: str,
                             checklist_id: str,
                             host: str | None = None,
                             tenant: str | None = None) -> dict:
    """Fetch checklist-level default settings (response sets, note placeholder, etc.).

    The checklist/get endpoint returns the checklist object which contains
    default settings that procedures inherit when they have no explicit settings.
    """
    host = host or HOST
    tenant = tenant or TENANT
    url = f"{host}/{tenant}/e/eng/{engagement_id}/api/v1.12.0/checklist/get"
    try:
        result = _api_post(session, url, {})
        for item in result:
            if item.get("id") == checklist_id:
                settings = item.get("settings") or {}
                if settings:
                    log.info("  Fetched checklist default settings")
                return settings
    except Exception as exc:
        log.debug("Could not fetch checklist defaults: %s", exc)
    return {}


def fetch_tag_lookup(session: http_lib.Session,
                     engagement_id: str,
                     host: str | None = None,
                     tenant: str | None = None) -> dict[str, str]:
    """Fetch all tags and return {tag_id: name} map.
    Used to resolve assertion IDs, area IDs, etc."""
    host = host or HOST
    tenant = tenant or TENANT
    url = f"{host}/{tenant}/e/eng/{engagement_id}/api/v1.12.0/tag/get"
    log.info("Fetching tags for ID resolution")
    try:
        tags = _api_post(session, url, {})
        return {t["id"]: t.get("name", "") for t in tags if "id" in t}
    except Exception as exc:
        log.warning("Could not fetch tags: %s", exc)
        return {}


# ── PROCEDURE HELPERS ────────────────────────────────────────────────────────

def _get_procedure_name(proc: dict) -> str:
    """Return the display name of a procedure."""
    summary = proc.get("summaryNames") or {}
    name = summary.get("en", "") or next(iter(summary.values()), "")
    if not name:
        name = strip_html(proc.get("text", ""))
    return name


def _get_procedure_display_text(proc: dict) -> str:
    """Return the full display text for a procedure, including number prefix."""
    name = _get_procedure_name(proc)
    number = proc.get("number", "")
    if number and name and not name.startswith(number):
        return f"{number}. {name}"
    return name


# ── PROCEDURE TREE ───────────────────────────────────────────────────────────

def build_procedure_tree(procedures: list[dict]) -> list[dict]:
    """Sort procedures into display order using parent-child tree + order field.
    Returns flat list in depth-first order."""
    by_id = {p["id"]: p for p in procedures if "id" in p}
    children_by_parent: dict[str, list[dict]] = {}
    roots: list[dict] = []

    for proc in procedures:
        pid = proc.get("parentId", "")
        if not pid or pid not in by_id:
            roots.append(proc)
        else:
            children_by_parent.setdefault(pid, []).append(proc)

    roots.sort(key=lambda p: p.get("order", ""))
    for children in children_by_parent.values():
        children.sort(key=lambda p: p.get("order", ""))

    result: list[dict] = []
    def walk(node):
        result.append(node)
        for child in children_by_parent.get(node.get("id", ""), []):
            walk(child)

    for root in roots:
        walk(root)

    return result


def _has_children(proc: dict, children_by_parent: dict[str, list]) -> bool:
    """Check if a procedure has child procedures."""
    return bool(children_by_parent.get(proc.get("id", "")))


def _has_response_sets(proc: dict) -> bool:
    """Check if a procedure has non-empty responseSets in its settings."""
    settings = proc.get("settings") or {}
    response_sets = settings.get("responseSets") or []
    if not response_sets:
        return False
    # Check that at least one response set has responses
    for rs in response_sets:
        if rs.get("responses"):
            return True
    return False


def classify_procedure(proc: dict, by_id: dict[str, dict],
                       children_by_parent: dict[str, list]) -> str:
    """Classify a procedure as section_header, group, procedure, or sub_procedure.

    Returns one of: 'section_header', 'group', 'procedure', 'sub_procedure'

    Uses the API 'type' field when available:
      - type='group' → section_header or group
      - type='procedure' → procedure, group (if has children + no RS), or sub_procedure
      - type='conclusion', 'taxabstract', etc. → procedure
    """
    api_type = proc.get("type", "")
    has_settings = _has_response_sets(proc)
    has_kids = _has_children(proc, children_by_parent)
    number = proc.get("number", "").strip()

    # API type=group → always section_header (these are organizational headings)
    if api_type == "group":
        return "section_header"

    # API type=procedure (or conclusion, taxabstract, etc.)
    if api_type in ("procedure", "conclusion", "taxabstract", ""):
        # Procedures with children and no response sets → group
        if has_kids and not has_settings:
            return "group"

        # Sub-procedures: lettered (a, b, c, etc.)
        if number and re.match(r'^[a-z]$', number):
            return "sub_procedure"

        # Regular procedures (even if no explicit responseSets — they inherit defaults)
        return "procedure"

    # Unknown types → default to procedure
    return "procedure"


# ── SETTINGS EXTRACTION ──────────────────────────────────────────────────────

def extract_procedure_settings(proc: dict, checklist_defaults: dict | None = None) -> dict:
    """Extract display settings from a procedure's settings object.
    Returns dict with keys matching Excel columns E-N.

    If the procedure has no explicit settings, falls back to checklist_defaults
    (fetched via checklist/get endpoint).

    The API uses different field names than the Excel output:
      - allowSignOffs (API) → Allow Sign offs (Excel)
      - allowNote (API) → Allow Input Notes in procedures (Excel)
      - notePlaceholder / notePlaceholders.en (API) → Notes Placeholder (Excel)
      - showResponsesBelow (API) → Show Response Beneath Procedure (Excel)
    """
    proc_settings = proc.get("settings") or {}
    defaults = checklist_defaults or {}

    # Determine if this procedure overrides checklist defaults
    has_own_settings = bool(proc_settings.get("responseSets"))
    override = has_own_settings

    # Use procedure settings if present, otherwise fall back to checklist defaults
    settings = proc_settings if has_own_settings else defaults

    allow_signoffs = settings.get("allowSignOffs",
                     settings.get("allowSignoffs", False))
    allow_notes = settings.get("allowNote",
                  settings.get("allowInputNotes", False))

    notes_placeholder = (settings.get("notePlaceholder", "")
                         or (settings.get("notePlaceholders") or {}).get("en", "")
                         or settings.get("notesPlaceholder", ""))
    allow_multiple_rows = settings.get("allowMultipleRows", False)
    show_beneath = settings.get("showResponsesBelow",
                   settings.get("showResponseBeneathProcedure", False))

    return {
        "override": override,
        "allow_signoffs": allow_signoffs,
        "allow_notes": allow_notes,
        "notes_placeholder": notes_placeholder,
        "allow_multiple_rows": allow_multiple_rows,
        "show_beneath": show_beneath,
    }


def get_response_set_rows(proc: dict, checklist_defaults: dict | None = None) -> list[dict]:
    """Return one dict per responseSet with response_placeholder,
    response_type, display_inline, and options.

    If the procedure has no explicit responseSets, falls back to
    checklist_defaults (fetched via checklist/get endpoint).

    API response set shape:
      type: "picklist" | "manual" | "multi-picklist"
      description / descriptions.en: placeholder text
      displayInline: boolean
      responses[]: {name, names.en, nonOptimal, ...}
    """
    proc_settings = proc.get("settings") or {}
    response_sets = proc_settings.get("responseSets") or []
    if not response_sets and checklist_defaults:
        response_sets = checklist_defaults.get("responseSets") or []
    if not response_sets:
        return [{}]
    settings = proc_settings if proc_settings.get("responseSets") else (checklist_defaults or {})

    rows = []
    for rs in response_sets:
        responses = rs.get("responses") or []
        options_parts = []
        for resp in responses:
            name = resp.get("name", "") or (resp.get("names") or {}).get("en", "")
            non_optimal = resp.get("nonOptimal", False)
            if name:
                label = f"{name} (Non optimal)" if non_optimal else name
                options_parts.append(label)

        # API uses 'type' on the responseSet, not 'responseType'
        resp_type_raw = (rs.get("type", "")
                         or rs.get("responseType", "")
                         or settings.get("responseType", ""))
        resp_type = {
            "picklist": "Picklist",
            "manual": "Manual",
            "multiPicklist": "Multi-Picklist",
            "multi-picklist": "Multi-Picklist",
        }.get(resp_type_raw, resp_type_raw)

        # Placeholder: API uses 'description' / 'descriptions.en', not 'placeholder'
        placeholder = (rs.get("description", "")
                       or (rs.get("descriptions") or {}).get("en", "")
                       or rs.get("placeholder", ""))

        rows.append({
            "response_placeholder": placeholder,
            "response_type": resp_type,
            "display_inline": rs.get("displayInline", True),
            "options": " | ".join(options_parts),
        })
    return rows if rows else [{}]


# ── STANDARDS & ASSERTIONS EXTRACTION ────────────────────────────────────────

def extract_standards(proc: dict) -> str:
    """Extract AICPA/CAS standard references from the procedure.

    Standards live in proc.attachables as objects with kind='citation'.
    The label is in labels.en or label field (e.g. 'AU-C 520.05').
    Sorted by order field for consistent output.
    """
    attachables = proc.get("attachables") or {}
    citations = []
    for att in attachables.values():
        if att.get("kind") == "citation":
            label = ((att.get("labels") or {}).get("en", "")
                     or att.get("label", ""))
            order = att.get("order", "")
            if label:
                citations.append((order, label.strip()))

    if not citations:
        # Fallback: try top-level fields
        for path in ("references", "standardReferences", "authoritativeReferences"):
            val = proc.get(path)
            if val and isinstance(val, list):
                parts = []
                for ref in val:
                    if isinstance(ref, dict):
                        name = (ref.get("name", "") or (ref.get("names") or {}).get("en", ""))
                        if name:
                            parts.append(strip_html(name))
                    elif isinstance(ref, str):
                        parts.append(ref)
                if parts:
                    return "\n".join(parts)
        return ""

    # Sort by order field and join with newlines
    citations.sort(key=lambda x: x[0])
    return "\n".join(c[1] for c in citations)


def extract_assertions(proc: dict, tag_lookup: dict[str, str]) -> str:
    """Extract the 5 author-selectable assertion abbreviations (C, E, A, V, PD).

    Authors select from exactly 5 assertions in the UI. The API stores 14
    "baseassertion" tags grouped under "Classes of Transactions" (CoT) and
    "Account Balances" (AB). Each UI assertion creates tags in both groups.

    Detection rules (verified empirically):
      C  = "Completeness" present (either group)
      E  = "Existence" present (AB group)
      A  = "Accuracy" (CoT) present  OR  "Accuracy, valuation and allocation" (AB) present
      V  = "Acc,val,alloc" (AB) present AND "Accuracy" (CoT) NOT present
           OR all 6 AB tags present (meaning all 5 assertions were selected)
      PD = "Presentation" present (either group)
    """
    tagging = proc.get("tagging") or {}
    base_assertions = tagging.get("baseassertion") or {}
    if not base_assertions:
        return ""

    # Collect all baseassertion names, tracking parent group
    all_names: set[str] = set()
    ab_names: set[str] = set()   # Account Balances group only
    for tag_ids in base_assertions.values():
        for tid in tag_ids:
            tag = tag_lookup.get(tid, "")
            if not tag:
                continue
            all_names.add(tag)

    # We need parent info to distinguish CoT vs AB for the A/V disambiguation.
    # Rebuild with parent awareness using the full tag objects from tag_lookup.
    # tag_lookup is {id: name}, but we also need parent info.
    # Instead, use a simpler heuristic: "Accuracy" (exact) = CoT tag,
    # "Accuracy, valuation and allocation" = AB tag. These names are unique.
    has_accuracy_cot = "Accuracy" in all_names
    has_ava_ab = "Accuracy, valuation and allocation" in all_names

    # Count AB tags by identifying the known AB baseassertion names
    _AB_NAMES = {"Existence", "Completeness", "Rights and obligations",
                 "Accuracy, valuation and allocation", "Classification", "Presentation"}
    for tag_ids in base_assertions.values():
        for tid in tag_ids:
            tag = tag_lookup.get(tid, "")
            if tag in _AB_NAMES:
                ab_names.add(tag)
    ab_full = len(ab_names) == 6  # all 6 AB baseassertions present

    result = []
    if "Completeness" in all_names:
        result.append("C")
    if "Existence" in all_names:
        result.append("E")
    if has_accuracy_cot or has_ava_ab:
        result.append("A")
    if has_ava_ab and (not has_accuracy_cot or ab_full):
        result.append("V")
    if "Presentation" in all_names:
        result.append("PD")

    return ", ".join(result)


def extract_guidance(proc: dict) -> str:
    """Extract lightbulb guidance text from the procedure."""
    guidances = proc.get("guidances") or {}
    guidance_html = guidances.get("en", "") or proc.get("guidance", "") or ""
    return strip_html(guidance_html)


# ── ID RESOLUTION ────────────────────────────────────────────────────────────

def _resolve(lookup: dict, id_obj) -> str:
    """Resolve an id-object ({id: ..., authorId: ...}) to a human-readable name."""
    if not id_obj:
        return ""
    raw_id = id_obj.get("id", "") if isinstance(id_obj, dict) else str(id_obj)
    return lookup.get(raw_id, raw_id[:12])


# Known customOrganizationTypeId → human-readable label mapping
_ORG_TYPE_LABELS: dict[str, str] = {
    "CorporationControlledPrivateCorporation": "Canadian Controlled Private Corporation (CCPC)",
    "CorporationControlledPublicCorporation": "Corporation Controlled by a Public Corporation",
    "OtherPrivateCorporation": "Other Private Corporation",
    "PublicCompany": "Public Company",
    "Individual": "Individual",
    "Coownership": "Co-ownership",
    "GeneralPartnership": "General Partnership",
    "LimitedPartnership": "Limited Partnership",
    "LimitedLiabilityPartnership": "Limited Liability Partnership",
    "JointVenture": "Joint Venture",
    "Trust": "Trust",
    "Cooperative": "Cooperative",
    "PensionFunds": "Pension Funds",
    "RegisteredCharity": "Registered Charity",
    "NotForProfit": "Not For Profit",
    "Government": "Government",
    "LimitedLiabilityCompany": "Limited Liability Company (LLC)",
    "CCorporation": "C-Corporation",
    "SCorporation": "S-Corporation",
    "SoleProprietorship": "Sole Proprietorship",
    "Partnership": "Partnership",
    "NotForProfitPrivate": "Not for Profit - Private",
    "NotForProfitPublic": "Not for Profit - Public",
    "PublicFoundation": "Public Foundation",
    "PrivateFoundation": "Private Foundation",
}


def _split_pascal_case(s: str) -> str:
    """Split a PascalCase string into words."""
    return re.sub(r"(?<=[a-z])(?=[A-Z])|(?<=[A-Z])(?=[A-Z][a-z])", " ", s)


def _resolve_org_type(cond: dict) -> str:
    """Resolve an organization_type condition to a human-readable name."""
    custom_id = cond.get("customOrganizationTypeId", "")
    if custom_id:
        return _ORG_TYPE_LABELS.get(custom_id, _split_pascal_case(custom_id))
    org_type = cond.get("organizationType", "")
    if org_type:
        return _ORG_TYPE_LABELS.get(org_type, _split_pascal_case(org_type))
    return cond.get("id", "unknown")[:12]


def _collect_procedure_ids_from_cond(cond: dict, ids: set) -> None:
    """Recursively collect procedureId values from a condition."""
    pid = (cond.get("procedureId") or {}).get("id")
    if pid:
        ids.add(pid)
    for sub in cond.get("conditions") or []:
        _collect_procedure_ids_from_cond(sub, ids)


def _collect_checklist_ids_from_cond(cond: dict, ids: set) -> None:
    """Recursively collect checklistId values from a condition."""
    cid = (cond.get("checklistId") or {}).get("id")
    if cid:
        ids.add(cid)
    for sub in cond.get("conditions") or []:
        _collect_checklist_ids_from_cond(sub, ids)


def build_id_lookup(session: http_lib.Session,
                    engagement_id: str,
                    procedures: list[dict],
                    doc_labels: dict[str, str],
                    host: str | None = None,
                    tenant: str | None = None) -> dict[str, str]:
    """Build a {id: human_name} lookup for all IDs referenced in visibility conditions.

    Resolves checklist names, procedure names, and response option names.
    """
    # Collect all unique IDs from visibility conditions
    procedure_ids: set = set()
    checklist_ids: set = set()
    for proc in procedures:
        for cond in (proc.get("visibility") or {}).get("conditions") or []:
            _collect_procedure_ids_from_cond(cond, procedure_ids)
            _collect_checklist_ids_from_cond(cond, checklist_ids)

    log.info("  %d procedure IDs, %d checklist IDs to resolve",
             len(procedure_ids), len(checklist_ids))

    # Seed with document labels
    lookup: dict[str, str] = dict(doc_labels)

    # Resolve checklist names
    for cid in checklist_ids:
        if cid in lookup:
            continue
        name = fetch_checklist_name(session, engagement_id, cid, host=host, tenant=tenant)
        if name:
            lookup[cid] = name
            log.debug("  Checklist %s → '%s'", cid[:12], name)

    # Resolve procedure names + response options
    for idx, pid in enumerate(procedure_ids, start=1):
        if pid in lookup:
            continue
        fetched = fetch_procedure_by_id(session, engagement_id, pid, host=host, tenant=tenant)
        if not fetched:
            continue
        proc_name = _get_procedure_name(fetched)
        if proc_name:
            lookup[pid] = proc_name
        # Response options from settings.responseSets
        for rs in (fetched.get("settings") or {}).get("responseSets") or []:
            for resp_opt in rs.get("responses") or []:
                rid  = resp_opt.get("id", "")
                name = resp_opt.get("name", "") or (resp_opt.get("names") or {}).get("en", "")
                if rid and name:
                    lookup[rid] = name
        if idx % 10 == 0:
            log.info("    ... %d / %d procedures resolved", idx, len(procedure_ids))

    log.info("  %d names resolved in lookup", len(lookup))
    return lookup


# ── VISIBILITY FORMATTING ────────────────────────────────────────────────────

def _format_response_condition(cond: dict, lookup: dict) -> str:
    """Format a single response condition as human-readable text.
    Output: '{checklist_name}\n{procedure_name}\n= {response}'
    """
    checklist = _resolve(lookup, cond.get("checklistId"))
    procedure = _resolve(lookup, cond.get("procedureId"))
    response  = _resolve(lookup, cond.get("responseId"))
    parts = []
    if checklist:
        parts.append(checklist)
    if procedure:
        parts.append(procedure)
    if response:
        parts.append(f"= {response.upper()}" if response.lower() in ("yes", "no", "true", "false")
                      else f"= {response}")
    return "\n".join(parts)


def _format_rmm_rank_condition(cond: dict, tag_lookup: dict[str, str]) -> str:
    """Format an rmm_rank visibility condition.

    API shape:
      type: "rmm_rank"
      tagId: {id: "..."} — the financial statement area tag
      assertionIds: ["id1", "id2", ...] — baseassertion tag IDs
      rmm: "low" | "medium" | "high"
      operator: "ge" | "gt" | "eq"

    Output examples:
      "Show when ANY of the following assertions of Cash and cash equivalents
       have RMM >= Medium\n\nAssertions include: completeness, existence, or accuracy"
      "Show when the completeness assertion of Cash and cash equivalents has RMM >= High"
    """
    # Resolve area name from tagId
    tag_id_obj = cond.get("tagId") or {}
    area_id = tag_id_obj.get("id", "") if isinstance(tag_id_obj, dict) else str(tag_id_obj)
    area_name = tag_lookup.get(area_id, area_id[:12])

    # Resolve assertion names
    assertion_ids = cond.get("assertionIds") or []
    assertion_names = []
    for aid in assertion_ids:
        name = tag_lookup.get(aid, "")
        if name:
            assertion_names.append(name.lower())

    # RMM level and operator
    rmm = (cond.get("rmm", "") or "").capitalize()
    operator = cond.get("operator", "ge")
    op_symbol = {"ge": ">=", "gt": ">", "eq": "=", "le": "<=", "lt": "<"}.get(operator, operator)

    if len(assertion_names) == 1:
        return (f"the {assertion_names[0]} assertion of "
                f"{area_name} has RMM {op_symbol}{rmm}")
    elif len(assertion_names) > 1:
        if len(assertion_names) == 2:
            joined = f"{assertion_names[0]} or {assertion_names[1]}"
        else:
            joined = ", ".join(assertion_names[:-1]) + ", or " + assertion_names[-1]
        return (f"ANY of the following assertions of {area_name} "
                f"have RMM {op_symbol} {rmm}\n\n"
                f"Assertions include: {joined}")
    else:
        return f"{area_name} has RMM {op_symbol} {rmm}"


def _format_single_condition(cond: dict, lookup: dict,
                             tag_lookup: dict[str, str] | None = None) -> str:
    """Format a single condition of any type."""
    ctype = cond.get("type", "")
    tag_lookup = tag_lookup or {}

    if ctype == "response":
        return _format_response_condition(cond, lookup)

    elif ctype == "rmm_rank":
        return _format_rmm_rank_condition(cond, tag_lookup)

    elif ctype == "condition_group":
        nested = cond.get("conditions") or []
        parts = []
        for sub in nested:
            parts.append(_format_single_condition(sub, lookup, tag_lookup))
        return "\n\n".join(parts)

    elif ctype == "organization_type":
        org_name = _resolve_org_type(cond)
        return f"Organization type = {org_name}"

    elif ctype == "consolidation":
        consolidated = cond.get("consolidated", False)
        return "Consolidated" if consolidated else "Not consolidated"

    elif ctype == "enum_value":
        # Visibility form condition: key = "VISIBILITYFORM.{areaTagId}", value = "testing" etc.
        key = cond.get("key", "")
        value = cond.get("conditionValue", "")
        # Resolve area tag from key
        parts = key.split(".", 1)
        if len(parts) == 2 and tag_lookup:
            area_name = tag_lookup.get(parts[1], parts[1][:12])
            return f"{parts[0]}: {area_name} = {value}"
        return f"{key} = {value}"

    elif ctype == "boolean_value":
        # Accounting estimate or other boolean flag: key = "ACCOUNTINGEST.SigEstCash"
        key = cond.get("key", "")
        value = cond.get("conditionValue", False)
        return f"{key} = {'TRUE' if value else 'FALSE'}"

    else:
        # Unknown condition type — show raw for discovery
        return f"[{ctype}] {json.dumps(cond)[:200]}"


def _parent_has_conditions(proc: dict, by_id: dict[str, dict]) -> bool:
    """Check if any ancestor has visibility conditions."""
    pid = proc.get("parentId", "")
    visited = {proc.get("id", "")}
    while pid and pid in by_id:
        if pid in visited:
            break
        visited.add(pid)
        parent = by_id[pid]
        if (parent.get("visibility") or {}).get("conditions"):
            return True
        pid = parent.get("parentId", "")
    return False


def format_visibility_columns(proc: dict, by_id: dict[str, dict],
                              lookup: dict[str, str],
                              tag_lookup: dict[str, str] | None = None) -> list[str]:
    """Format visibility conditions into up to 5 strings for columns P-T.

    Returns list of 5 strings (empty string for unused columns).
    """
    vis = proc.get("visibility") or {}
    conditions = vis.get("conditions") or []
    normally_visible = vis.get("normallyVisible", True)
    all_needed = vis.get("allConditionsNeeded", False)
    tag_lookup = tag_lookup or {}

    # No conditions at all
    if not conditions:
        if _parent_has_conditions(proc, by_id):
            return ["Inherited from above.", "", "", "", ""]
        return ["None", "", "", "", ""]

    # Determine direction
    direction = "Show" if not normally_visible else "Hide"

    # Single condition
    if len(conditions) == 1:
        cond = conditions[0]
        body = _format_single_condition(cond, lookup, tag_lookup)
        text = f"{direction} when:\n{body}"
        return [text, "", "", "", ""]

    # Multiple conditions
    if all_needed:
        # AND logic — each condition in its own column
        columns = []
        for i, cond in enumerate(conditions):
            body = _format_single_condition(cond, lookup, tag_lookup)
            if i == 0:
                text = f"{direction} when ALL are met:\n\n{body}\n"
            else:
                text = f"AND when {body}"
            columns.append(text)
        while len(columns) < 5:
            columns.append("")
        return columns[:5]
    else:
        # OR logic — group rmm_rank conditions together, separate others
        rmm_conditions = [c for c in conditions if c.get("type") == "rmm_rank"]
        other_conditions = [c for c in conditions if c.get("type") != "rmm_rank"]

        if rmm_conditions and not other_conditions:
            # All conditions are rmm_rank — merge assertion lists
            # They typically share the same area and rmm level
            first = rmm_conditions[0]
            tag_id_obj = first.get("tagId") or {}
            area_id = tag_id_obj.get("id", "") if isinstance(tag_id_obj, dict) else str(tag_id_obj)
            area_name = tag_lookup.get(area_id, area_id[:12])
            rmm = (first.get("rmm", "") or "").capitalize()
            operator = first.get("operator", "ge")
            op_symbol = {"ge": ">=", "gt": ">", "eq": "=", "le": "<=", "lt": "<"}.get(operator, operator)

            all_assertion_names = []
            seen_names = set()
            for cond in rmm_conditions:
                for aid in cond.get("assertionIds") or []:
                    name = tag_lookup.get(aid, "")
                    if name and name.lower() not in seen_names:
                        all_assertion_names.append(name.lower())
                        seen_names.add(name.lower())

            if len(all_assertion_names) == 1:
                text = (f"{direction} when the {all_assertion_names[0]} assertion of "
                        f" {area_name} has RMM {op_symbol}{rmm}")
            elif len(all_assertion_names) > 1:
                if len(all_assertion_names) == 2:
                    joined = f"{all_assertion_names[0]} or {all_assertion_names[1]}"
                else:
                    joined = ", ".join(all_assertion_names[:-1]) + ", or " + all_assertion_names[-1]
                text = (f"{direction} when ANY of the following assertions of "
                        f"{area_name} have RMM {op_symbol} {rmm}\n\n"
                        f"Assertions include: {joined}")
            else:
                text = f"{direction} when {area_name} has RMM {op_symbol} {rmm}"
            return [text, "", "", "", ""]

        # Mixed condition types
        body_parts = []
        for cond in conditions:
            body_parts.append(_format_single_condition(cond, lookup, tag_lookup))
        text = f"{direction} when ANY:\n\n" + "\n\n".join(body_parts)
        return [text, "", "", "", ""]


# ── EXCEL STYLES ─────────────────────────────────────────────────────────────

HEADER_FILL    = PatternFill("solid", fgColor="FF073763")    # dark blue
HEADER_FONT    = Font(bold=True, color="FFFFFFFF", size=10)  # white bold
HEADER_ALIGN   = Alignment(horizontal="center", vertical="top", wrap_text=True)

SECTION_FILL   = PatternFill("solid", fgColor="FF9FC5E8")    # light blue
SECTION_FONT   = Font(bold=True, color="FF073763", size=10)  # dark blue bold

DATA_FONT      = Font(size=10)
DATA_ALIGN     = Alignment(vertical="top", wrap_text=True)
DATA_ALIGN_NW  = Alignment(vertical="top", wrap_text=False)

# Column layout: A through S (19 columns)
COLUMN_HEADERS = [
    "Procedure Text",                    # A
    "Authoritative Reference",           # B
    "Assertions",                        # C
    "Lightbulb Guidance",                # D
    "Response Placeholder",              # E
    "Response Type",                     # F
    "Options",                           # G
    "Display inline",                    # H
    "Allow Input Notes in procedures",   # I
    "Notes Placeholder",                 # J
    "Allow Sign offs",                   # K
    "Condition 1",                       # L
    "Condition 2",                       # M
    "Condition 3",                       # N
    "Condition 4",                       # O
    "Condition 5",                       # P
    "Allow multiple rows",               # Q
    "Show Response Beneath Procedure",   # R
    "Override the checklist settings",   # S
]

COLUMN_WIDTHS = {
    "A": 80, "B": 18, "C": 18, "D": 50,
    "E": 22, "F": 15, "G": 40, "H": 12,
    "I": 12, "J": 22, "K": 12,
    "L": 50, "M": 50, "N": 50, "O": 50, "P": 50,
    "Q": 12, "R": 12, "S": 12,
}


# ── EXCEL WRITING ────────────────────────────────────────────────────────────

def _apply_header_style(ws, row_num: int, max_col: int = 19):
    """Apply header fill/font to an entire row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN


def _apply_section_style(ws, row_num: int, max_col: int = 19):
    """Apply section header fill/font to an entire row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = SECTION_FILL
        if col == 1:
            cell.font = SECTION_FONT
        cell.alignment = DATA_ALIGN


def _write_procedure_row(ws, row_num: int, text: str, standards: str,
                         assertions: str, guidance: str, settings: dict,
                         vis_columns: list[str], rs_row: dict):
    """Write a single procedure row with all columns."""
    ws.cell(row=row_num, column=1, value=text).alignment = DATA_ALIGN           # A
    ws.cell(row=row_num, column=2, value=standards).alignment = DATA_ALIGN       # B
    ws.cell(row=row_num, column=3, value=assertions).alignment = DATA_ALIGN_NW   # C
    ws.cell(row=row_num, column=4, value=guidance).alignment = DATA_ALIGN        # D

    # Response set columns E-H
    ws.cell(row=row_num, column=5, value=rs_row.get("response_placeholder", "")).alignment = DATA_ALIGN  # E
    ws.cell(row=row_num, column=6, value=rs_row.get("response_type", "")).alignment = DATA_ALIGN_NW  # F
    ws.cell(row=row_num, column=7, value=rs_row.get("options", "")).alignment = DATA_ALIGN  # G
    ws.cell(row=row_num, column=8, value=rs_row.get("display_inline")).alignment = DATA_ALIGN_NW  # H

    # Settings columns I-K
    ws.cell(row=row_num, column=9, value=settings.get("allow_notes")).alignment = DATA_ALIGN_NW  # I
    ws.cell(row=row_num, column=10, value=settings.get("notes_placeholder", "")).alignment = DATA_ALIGN  # J
    ws.cell(row=row_num, column=11, value=settings.get("allow_signoffs")).alignment = DATA_ALIGN_NW  # K

    # Visibility columns L-P (columns 12-16)
    for i, vis_text in enumerate(vis_columns):
        ws.cell(row=row_num, column=12 + i, value=vis_text).alignment = DATA_ALIGN  # L-P

    # Settings columns Q-S
    ws.cell(row=row_num, column=17, value=settings.get("allow_multiple_rows")).alignment = DATA_ALIGN_NW  # Q
    ws.cell(row=row_num, column=18, value=settings.get("show_beneath")).alignment = DATA_ALIGN_NW  # R
    ws.cell(row=row_num, column=19, value=settings.get("override")).alignment = DATA_ALIGN_NW  # S


def _write_response_set_only_row(ws, row_num: int, rs_row: dict):
    """Write a row that only has response set columns E-H (for multi-response procedures)."""
    ws.cell(row=row_num, column=5, value=rs_row.get("response_placeholder", "")).alignment = DATA_ALIGN
    ws.cell(row=row_num, column=6, value=rs_row.get("response_type", "")).alignment = DATA_ALIGN_NW
    ws.cell(row=row_num, column=7, value=rs_row.get("options", "")).alignment = DATA_ALIGN
    ws.cell(row=row_num, column=8, value=rs_row.get("display_inline")).alignment = DATA_ALIGN_NW


def build_checklist_sheet(ws, procedures: list[dict], lookup: dict[str, str],
                          tag_lookup: dict[str, str] | None = None,
                          checklist_defaults: dict | None = None):
    """Write a single checklist's data to a worksheet."""
    tag_lookup = tag_lookup or {}
    checklist_defaults = checklist_defaults or {}

    # ── Row 1: Category headers ──
    # Style entire row first
    _apply_header_style(ws, 1)

    ws.merge_cells("E1:K1")
    ws["E1"].value = "Cloud Settings"
    ws["E1"].fill = HEADER_FILL
    ws["E1"].font = HEADER_FONT
    ws["E1"].alignment = HEADER_ALIGN

    ws.merge_cells("L1:P1")
    ws["L1"].value = "Visibility Settings"
    ws["L1"].fill = HEADER_FILL
    ws["L1"].font = HEADER_FONT
    ws["L1"].alignment = HEADER_ALIGN

    ws.merge_cells("Q1:S1")
    ws["Q1"].value = "Cloud Settings"
    ws["Q1"].fill = HEADER_FILL
    ws["Q1"].font = HEADER_FONT
    ws["Q1"].alignment = HEADER_ALIGN

    # ── Row 2: Column headers ──
    for col_idx, header in enumerate(COLUMN_HEADERS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN

    # Set column widths
    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze panes at A3
    ws.freeze_panes = "A3"

    # Auto-filter on row 2
    ws.auto_filter.ref = f"A2:S2"

    # ── Build tree and classify ──
    by_id = {p["id"]: p for p in procedures if "id" in p}
    children_by_parent: dict[str, list[dict]] = {}
    for proc in procedures:
        pid = proc.get("parentId", "")
        if pid and pid in by_id:
            children_by_parent.setdefault(pid, []).append(proc)

    ordered = build_procedure_tree(procedures)

    # ── Write data rows ──
    row_num = 3
    for proc in ordered:
        proc_type = classify_procedure(proc, by_id, children_by_parent)
        text = _get_procedure_display_text(proc)

        if proc_type == "section_header":
            ws.cell(row=row_num, column=1, value=text)
            _apply_section_style(ws, row_num)
            # Section headers can have guidance (lightbulb)
            guidance_sh = extract_guidance(proc)
            if guidance_sh:
                ws.cell(row=row_num, column=4, value=guidance_sh).alignment = DATA_ALIGN
            # Section headers can still have visibility conditions
            vis_columns_sh = format_visibility_columns(proc, by_id, lookup, tag_lookup)
            for i, vis_text in enumerate(vis_columns_sh):
                ws.cell(row=row_num, column=12 + i, value=vis_text).alignment = DATA_ALIGN
            row_num += 1
            continue

        # Extract all data
        settings = extract_procedure_settings(proc, checklist_defaults)
        standards = extract_standards(proc)
        assertions = extract_assertions(proc, tag_lookup)
        guidance = extract_guidance(proc)
        vis_columns = format_visibility_columns(proc, by_id, lookup, tag_lookup)
        response_rows = get_response_set_rows(proc, checklist_defaults)

        if proc_type == "group":
            # Group rows: text + optional standards/assertions + visibility
            standards_g = extract_standards(proc)
            assertions_g = extract_assertions(proc, tag_lookup)
            vis_columns_g = format_visibility_columns(proc, by_id, lookup, tag_lookup)
            ws.cell(row=row_num, column=1, value=text).alignment = DATA_ALIGN
            ws.cell(row=row_num, column=2, value=standards_g).alignment = DATA_ALIGN
            ws.cell(row=row_num, column=3, value=assertions_g).alignment = DATA_ALIGN_NW
            for i, vis_text in enumerate(vis_columns_g):
                ws.cell(row=row_num, column=12 + i, value=vis_text).alignment = DATA_ALIGN
            row_num += 1
            continue

        # Full procedure or sub-procedure
        if len(response_rows) <= 1:
            _write_procedure_row(ws, row_num, text, standards, assertions,
                                 guidance, settings, vis_columns,
                                 response_rows[0] if response_rows else {})
            row_num += 1
        else:
            # Multi-response-set: write first row with all data
            start_row = row_num
            _write_procedure_row(ws, row_num, text, standards, assertions,
                                 guidance, settings, vis_columns, response_rows[0])
            row_num += 1

            # Write subsequent rows with only response set columns
            for rs_row in response_rows[1:]:
                _write_response_set_only_row(ws, row_num, rs_row)
                row_num += 1

            end_row = row_num - 1

            # Merge cells that span all response rows
            if start_row < end_row:
                merge_cols = ["A", "B", "C", "D",
                              "J", "K", "L", "M", "N", "O", "P",
                              "Q", "R", "S"]
                for col in merge_cols:
                    ws.merge_cells(f"{col}{start_row}:{col}{end_row}")


def _sanitize_sheet_name(name: str) -> str:
    """Truncate and sanitize a sheet name for Excel (max 31 chars)."""
    name = re.sub(r'[\\/?*\[\]:"]', '', name).strip()
    if len(name) > 31:
        name = name[:28] + "..."
    return name or "Sheet"


def _unique_sheet_name(name: str, existing: set[str]) -> str:
    """Ensure a sheet name is unique by appending a suffix if needed."""
    base = name
    counter = 2
    while name in existing:
        suffix = f" ({counter})"
        name = base[:31 - len(suffix)] + suffix
        counter += 1
    existing.add(name)
    return name


# ── ORCHESTRATOR ─────────────────────────────────────────────────────────────

def generate_report_bytes(
    engagement_id: str,
    document_id: str = "",
    host: str | None = None,
    tenant: str | None = None,
) -> bytes:
    """Generate checklist extraction Excel and return bytes.

    If document_id is provided, extract only that checklist.
    Otherwise, extract all checklists in the engagement.
    """
    host = host or HOST
    tenant = tenant or TENANT
    env_prefix = _env_prefix_from_host(host)

    log.info("Generating checklist report for engagement %s", engagement_id)

    session = make_session(env_prefix, host=host, tenant=tenant)

    # Fetch all documents
    documents = fetch_documents(session, engagement_id, host=host, tenant=tenant)
    log.info("Fetched %d documents in engagement", len(documents))

    # Fetch tags for assertion/area resolution
    tag_lookup = fetch_tag_lookup(session, engagement_id, host=host, tenant=tenant)

    # Build document label lookup (index by both id and content field)
    doc_labels: dict[str, str] = {}
    # Also build a mapping from document id → content field (the real checklistId)
    doc_id_to_content: dict[str, str] = {}
    for doc in documents:
        did     = doc.get("id", "")
        content = doc.get("content", "")
        number  = doc.get("number", "")
        names   = doc.get("names") or {}
        name    = names.get("en", "") or doc.get("name", "")
        label   = f"{number} {name}".strip() if number else name
        if did:
            doc_labels[did] = label
            if content:
                doc_id_to_content[did] = content
        if content:
            doc_labels[content] = label

    # Determine which checklists to extract
    if document_id:
        # Single checklist mode — resolve document_id to its content field
        # (procedures use the content field as their checklistId)
        checklist_id = doc_id_to_content.get(document_id, document_id)
        label = doc_labels.get(document_id, doc_labels.get(checklist_id, "Checklist"))
        checklists_to_extract = [(checklist_id, label)]
    else:
        # Engagement mode — try each document using its content field
        checklists_to_extract = []
        for doc in documents:
            content = doc.get("content", "") or doc.get("id", "")
            label = doc_labels.get(content, content)
            checklists_to_extract.append((content, label))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    existing_names: set[str] = set()
    sheets_created = 0

    for idx, (checklist_id, label) in enumerate(checklists_to_extract, start=1):
        log.info("Processing %d/%d: %s", idx, len(checklists_to_extract), label)
        try:
            procedures = fetch_procedures(session, engagement_id, checklist_id,
                                          host=host, tenant=tenant)
            if not procedures:
                log.debug("  No procedures in %s — skipping", label)
                continue

            log.info("  %d procedures in %s", len(procedures), label)

            # Build ID lookup for visibility resolution
            id_lookup = build_id_lookup(session, engagement_id, procedures,
                                        doc_labels, host=host, tenant=tenant)

            # Fetch checklist-level default settings
            cl_defaults = fetch_checklist_defaults(session, engagement_id,
                                                    checklist_id, host=host,
                                                    tenant=tenant)

            # Create sheet
            sheet_name = _sanitize_sheet_name(label)
            sheet_name = _unique_sheet_name(sheet_name, existing_names)
            ws = wb.create_sheet(title=sheet_name)
            build_checklist_sheet(ws, procedures, id_lookup, tag_lookup,
                                 cl_defaults)
            sheets_created += 1

        except Exception as exc:
            log.warning("Skipping %s: %s", label, exc)

    if sheets_created == 0:
        raise ValueError(
            "No checklists found in this engagement. "
            "The documents may not contain any checklist procedures."
        )

    log.info("Created %d checklist sheet(s)", sheets_created)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── MOCK DATA ────────────────────────────────────────────────────────────────

def _mock_procedures() -> list[dict]:
    """Sample checklist procedures covering all row types and visibility patterns."""
    return [
        # Section header: "Procedures"
        {"id": "sh1", "parentId": "", "checklistId": "cl1", "order": "a",
         "number": "", "summaryNames": {"en": "Procedures"},
         "text": "<p>Procedures</p>", "settings": {}},

        # Section header: "Assessed risks"
        {"id": "sh2", "parentId": "sh1", "checklistId": "cl1", "order": "a1",
         "number": "", "summaryNames": {"en": "Assessed risks"},
         "text": "<p>Assessed risks</p>", "settings": {}},

        # Procedure 1: with settings, visibility=None
        {"id": "p1", "parentId": "sh2", "checklistId": "cl1", "order": "a1a",
         "number": "1",
         "text": "<p>Review pervasive risk factors and related proposed audit responses in 4-200 Financial statement areas</p>",
         "summaryNames": {"en": "Review pervasive risk factors and related proposed audit responses in 4-200 Financial statement areas"},
         "settings": {
             "override": True, "allowSignoffs": True, "allowInputNotes": True,
             "notesPlaceholder": "Response and comments",
             "allowMultipleRows": False,
             "showResponseBeneathProcedure": False,
             "responseType": "picklist",
             "responseSets": [{"responseType": "picklist", "displayInline": True,
                               "responses": [{"id": "r1", "name": "Completed"}]}],
         },
         "visibility": {"normallyVisible": True, "conditions": []}},

        # Section header: "Planned response to assessed risks"
        {"id": "sh3", "parentId": "sh1", "checklistId": "cl1", "order": "a2",
         "number": "", "summaryNames": {"en": "Planned response to assessed risks at the assertion level"},
         "text": "<p>Planned response to assessed risks at the assertion level</p>",
         "settings": {}},

        # Procedure 4: with response-based visibility
        {"id": "p4", "parentId": "sh3", "checklistId": "cl1", "order": "a2a",
         "number": "4",
         "text": "<p>Will substantive analytical procedures be used to respond to assessed risks?</p>",
         "summaryNames": {"en": "Will substantive analytical procedures be used to respond to assessed risks?"},
         "settings": {
             "override": True, "allowSignoffs": True, "allowInputNotes": True,
             "notesPlaceholder": "Response and comments",
             "allowMultipleRows": False,
             "showResponseBeneathProcedure": False,
             "responseSets": [{"responseType": "picklist", "displayInline": True,
                               "responses": [{"id": "r-yes", "name": "Yes"},
                                             {"id": "r-no", "name": "No"}]}],
         },
         "visibility": {"normallyVisible": True, "conditions": []}},

        # Procedure 5: with visibility show-when condition
        {"id": "p5", "parentId": "sh3", "checklistId": "cl1", "order": "a2b",
         "number": "5",
         "text": "<p>Explain why and consider the impact on the audit report.</p>",
         "summaryNames": {"en": "Explain why and consider the impact on the audit report."},
         "settings": {
             "override": True, "allowSignoffs": True, "allowInputNotes": False,
             "allowMultipleRows": False,
             "showResponseBeneathProcedure": False,
             "responseSets": [{"responseType": "manual", "displayInline": True,
                               "placeholder": "Explanation and comments",
                               "responses": []}],
         },
         "visibility": {
             "normallyVisible": False,
             "allConditionsNeeded": False,
             "conditions": [{
                 "type": "response",
                 "checklistId": {"id": "cl1"},
                 "procedureId": {"id": "p4"},
                 "responseId": {"id": "r-no"},
             }],
         }},

        # Section header: "Procedures" (second group)
        {"id": "sh4", "parentId": "", "checklistId": "cl1", "order": "b",
         "number": "", "summaryNames": {"en": "Procedures"},
         "text": "<p>Procedures</p>", "settings": {}},

        # Section header: "Substantive procedures"
        {"id": "sh5", "parentId": "sh4", "checklistId": "cl1", "order": "b1",
         "number": "", "summaryNames": {"en": "Substantive procedures"},
         "text": "<p>Substantive procedures</p>", "settings": {}},

        # Procedure 8: with standards and assertions
        {"id": "p8", "parentId": "sh5", "checklistId": "cl1", "order": "b1a",
         "number": "8",
         "text": "<p>Analytical procedures\nDevelop and document expectations for the period-end cash balance</p>",
         "summaryNames": {"en": "Analytical procedures\nDevelop and document expectations for the period-end cash balance"},
         "references": [{"name": "AU-C 520.05"}],
         "assertions": [{"name": "Completeness"}, {"name": "Existence"},
                        {"name": "Accuracy"}, {"name": "Valuation"}],
         "settings": {
             "override": False, "allowSignoffs": True, "allowInputNotes": True,
             "notesPlaceholder": "Response and comments",
             "allowMultipleRows": False,
             "showResponseBeneathProcedure": False,
             "responseSets": [{"responseType": "picklist", "displayInline": True,
                               "responses": [
                                   {"id": "r-ok", "name": "Completed, no exceptions"},
                                   {"id": "r-ex", "name": "Completed with exceptions", "nonOptimal": True},
                               ]}],
         },
         "guidances": {"en": "<p>You can use the bank confirmation template provided A.130 Bank confirmation</p>"},
         "visibility": {"normallyVisible": True, "conditions": []}},

        # Group procedure 15: has children, no response sets, with visibility
        {"id": "g15", "parentId": "sh5", "checklistId": "cl1", "order": "b1b",
         "number": "15",
         "text": "<p>Unrecorded cash and bank balances</p>",
         "summaryNames": {"en": "Unrecorded cash and bank balances"},
         "assertions": [{"name": "Completeness"}],
         "settings": {},
         "visibility": {
             "normallyVisible": False,
             "allConditionsNeeded": False,
             "conditions": [{
                 "type": "response",
                 "checklistId": {"id": "cl1"},
                 "procedureId": {"id": "p-rmm-completeness"},
                 "responseId": {"id": "r-rmm-high"},
             }],
         }},

        # Sub-procedure a: inherits visibility from parent g15
        {"id": "sp15a", "parentId": "g15", "checklistId": "cl1", "order": "b1b1",
         "number": "a",
         "text": "<p>Ask personnel familiar with, or handling, cash transactions whether they are aware of any unrecorded bank balances</p>",
         "summaryNames": {"en": "Ask personnel familiar with, or handling, cash transactions whether they are aware of any unrecorded bank balances"},
         "settings": {
             "override": False, "allowSignoffs": True, "allowInputNotes": True,
             "notesPlaceholder": "Response and comments",
             "allowMultipleRows": False,
             "showResponseBeneathProcedure": False,
             "responseSets": [{"responseType": "picklist", "displayInline": True,
                               "responses": [
                                   {"id": "r-ok", "name": "Completed, no exceptions"},
                                   {"id": "r-ex", "name": "Completed with exceptions", "nonOptimal": True},
                               ]}],
         },
         "visibility": {"normallyVisible": True, "conditions": []}},

        # Sub-procedure b: inherits visibility
        {"id": "sp15b", "parentId": "g15", "checklistId": "cl1", "order": "b1b2",
         "number": "b",
         "text": "<p>Review correspondence or other evidence available to ensure that accounts used for operating are properly recorded</p>",
         "summaryNames": {"en": "Review correspondence or other evidence available to ensure that accounts used for operating are properly recorded"},
         "settings": {
             "override": False, "allowSignoffs": True, "allowInputNotes": True,
             "notesPlaceholder": "Response and comments",
             "allowMultipleRows": False,
             "showResponseBeneathProcedure": False,
             "responseSets": [{"responseType": "picklist", "displayInline": True,
                               "responses": [
                                   {"id": "r-ok", "name": "Completed, no exceptions"},
                                   {"id": "r-ex", "name": "Completed with exceptions", "nonOptimal": True},
                               ]}],
         },
         "visibility": {"normallyVisible": True, "conditions": []}},

        # Procedure with multi-condition AND visibility
        {"id": "p24", "parentId": "sh5", "checklistId": "cl1", "order": "b1c",
         "number": "24",
         "text": "<p>Component entities\nIdentify and document transfers between bank accounts</p>",
         "summaryNames": {"en": "Component entities\nIdentify and document transfers between bank accounts"},
         "settings": {
             "override": False, "allowSignoffs": True, "allowInputNotes": True,
             "notesPlaceholder": "Response and comments",
             "allowMultipleRows": False,
             "showResponseBeneathProcedure": False,
             "responseSets": [{"responseType": "picklist", "displayInline": True,
                               "responses": [
                                   {"id": "r-ok", "name": "Completed, no exceptions"},
                                   {"id": "r-ex", "name": "Completed with exceptions", "nonOptimal": True},
                               ]}],
         },
         "visibility": {
             "normallyVisible": False,
             "allConditionsNeeded": True,
             "conditions": [
                 {
                     "type": "response",
                     "checklistId": {"id": "cl-setup"},
                     "procedureId": {"id": "p-component"},
                     "responseId": {"id": "r-yes"},
                 },
                 {
                     "type": "response",
                     "checklistId": {"id": "cl1"},
                     "procedureId": {"id": "p-rmm-accuracy"},
                     "responseId": {"id": "r-rmm-high"},
                 },
             ],
         }},

        # Procedure 27: multiple response sets
        {"id": "p27", "parentId": "sh5", "checklistId": "cl1", "order": "b1d",
         "number": "27",
         "text": "<p>Information\nNature of estimate(s) addressed by this worksheet</p>",
         "summaryNames": {"en": "Information\nNature of estimate(s) addressed by this worksheet"},
         "settings": {
             "override": True, "allowSignoffs": True, "allowInputNotes": True,
             "notesPlaceholder": "Response and comments",
             "allowMultipleRows": False,
             "showResponseBeneathProcedure": False,
             "responseSets": [
                 {"responseType": "manual", "displayInline": True,
                  "placeholder": "Nature of estimate(s)",
                  "responses": []},
                 {"responseType": "multiPicklist", "displayInline": True,
                  "placeholder": "Assertions affected",
                  "responses": [
                      {"id": "r-c", "name": "Completeness"},
                      {"id": "r-e", "name": "Existence"},
                      {"id": "r-a", "name": "Accuracy"},
                      {"id": "r-v", "name": "Valuation"},
                      {"id": "r-pd", "name": "Presentation and Disclosure"},
                  ]},
             ],
         },
         "visibility": {"normallyVisible": True, "conditions": []}},
    ]


def _mock_id_lookup() -> dict[str, str]:
    """Mock ID lookup for visibility resolution."""
    return {
        "cl1": "A.100 Cash and cash equivalents - Audit procedures",
        "cl-setup": "1-200 Engagement set up",
        "p4": "Will substantive analytical procedures be used to respond to assessed risks?",
        "p-rmm-completeness": "Completeness assertion RMM",
        "p-rmm-accuracy": "Accuracy assertion RMM",
        "p-component": "Are there transactions between component entities?",
        "r-no": "No",
        "r-yes": "YES",
        "r-rmm-high": "High",
    }


def run_mock():
    """Generate a report from mock data for testing."""
    procedures = _mock_procedures()
    lookup = _mock_id_lookup()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title="A.100 Cash - Audit proce...")
    build_checklist_sheet(ws, procedures, lookup, {})

    os.makedirs(os.path.dirname(OUTPUT_FILE) or ".", exist_ok=True)
    wb.save(OUTPUT_FILE)
    log.info("Mock report written to %s (%d procedures)", OUTPUT_FILE, len(procedures))


# ── DISCOVER MODE ────────────────────────────────────────────────────────────

def run_discover(engagement_id: str, checklist_id: str = "",
                 host: str | None = None, tenant: str | None = None):
    """Dump raw procedure JSON for field path discovery."""
    host = host or HOST
    tenant = tenant or TENANT
    env_prefix = _env_prefix_from_host(host)
    session = make_session(env_prefix, host=host, tenant=tenant)

    if not checklist_id:
        # Find the first checklist with procedures
        documents = fetch_documents(session, engagement_id, host=host, tenant=tenant)
        for doc in documents:
            cid = doc.get("content", "") or doc.get("id", "")
            procs = fetch_procedures(session, engagement_id, cid, host=host, tenant=tenant)
            if procs:
                checklist_id = cid
                log.info("Found checklist with %d procedures: %s",
                         len(procs), doc.get("names", {}).get("en", cid))
                break

    if not checklist_id:
        log.error("No checklists with procedures found.")
        return

    procedures = fetch_procedures(session, engagement_id, checklist_id,
                                  host=host, tenant=tenant)
    log.info("Fetched %d procedures", len(procedures))

    # Dump first 3 procedures (or fewer)
    for i, proc in enumerate(procedures[:3]):
        print(f"\n{'='*60}")
        print(f"PROCEDURE {i+1}: {_get_procedure_name(proc)}")
        print(f"{'='*60}")
        print(json.dumps(proc, indent=2, ensure_ascii=False))

    # Also dump one with visibility conditions if available
    for proc in procedures:
        vis = proc.get("visibility") or {}
        if vis.get("conditions"):
            print(f"\n{'='*60}")
            print(f"PROCEDURE WITH VISIBILITY: {_get_procedure_name(proc)}")
            print(f"{'='*60}")
            print(json.dumps(proc, indent=2, ensure_ascii=False))
            break

    # Dump all unique top-level keys across all procedures
    all_keys: set = set()
    for proc in procedures:
        all_keys.update(proc.keys())
    print(f"\n{'='*60}")
    print(f"ALL UNIQUE TOP-LEVEL KEYS ({len(all_keys)}):")
    print(f"{'='*60}")
    for key in sorted(all_keys):
        print(f"  {key}")


# ── CLI ───────────────────────────────────────────────────────────────────────

CW_URL_PATTERN = re.compile(
    r"https?://([^/]+)/([^/]+)/e/eng/([^/]+)"
)
CW_DOC_PATTERN = re.compile(
    r"#/(?:efinancials|checklist)/([^/?\s]+)"
)


def main():
    from dotenv import load_dotenv
    load_dotenv(os.path.join(os.path.dirname(__file__), "..", ".env"))

    parser = argparse.ArgumentParser(
        description="Extract checklist content from a Caseware Cloud SE author template."
    )
    parser.add_argument("--mock", action="store_true",
                        help="Generate report from mock data (no API calls)")
    parser.add_argument("--url", type=str, default="",
                        help="Caseware engagement or checklist URL")
    parser.add_argument("--output", type=str, default=OUTPUT_FILE,
                        help="Output Excel file path")
    parser.add_argument("--discover", action="store_true",
                        help="Dump raw procedure JSON for field path discovery")
    args = parser.parse_args()

    if args.mock:
        run_mock()
        return

    if not args.url:
        parser.error("Provide --url or use --mock for testing.")

    match = CW_URL_PATTERN.search(args.url)
    if not match:
        parser.error("Invalid Caseware URL format.")

    host = f"https://{match.group(1)}"
    tenant = match.group(2)
    engagement_id = match.group(3)

    doc_match = CW_DOC_PATTERN.search(args.url)
    document_id = doc_match.group(1) if doc_match else ""

    if args.discover:
        run_discover(engagement_id, document_id, host=host, tenant=tenant)
        return

    excel_bytes = generate_report_bytes(
        engagement_id=engagement_id,
        document_id=document_id,
        host=host,
        tenant=tenant,
    )

    os.makedirs(os.path.dirname(args.output) or ".", exist_ok=True)
    with open(args.output, "wb") as f:
        f.write(excel_bytes)
    log.info("Report written to %s", args.output)


if __name__ == "__main__":
    main()
